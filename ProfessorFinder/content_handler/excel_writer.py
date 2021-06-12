import os
import openpyxl


def check():
    dst = 'professor_info'
    if not os.path.isdir(dst):
        os.mkdir(dst)


def write_excel(uni_departments):
    university = tuple(uni_departments.keys())[0]
    filename = 'professor_info/' + university + '.xlsx'
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    for department, ok in uni_departments.items():
        if ok:
            department = department()
            print('processing: ' + university + department.name)
            if department.name in wb.sheetnames:
                wb.remove(wb[department.name])
            ws = wb.create_sheet(department.name, 0)
            ws.append(['学校', '院系', '姓名', '邮箱', '主页'])
            for each in department.run():
                ws.append(each)
    wb.save(filename)
    wb.close()
